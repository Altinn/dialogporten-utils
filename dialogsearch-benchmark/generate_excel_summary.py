#!/usr/bin/env python3
import argparse
import csv
import math
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Union

try:
    from stats_utils import summarize
except ModuleNotFoundError:
    def percentile(values: List[float], pct: float) -> float:
        if not values:
            return math.nan
        ordered = sorted(values)
        if pct <= 0:
            return ordered[0]
        if pct >= 100:
            return ordered[-1]
        rank = (pct / 100) * (len(ordered) - 1)
        low = int(math.floor(rank))
        high = int(math.ceil(rank))
        if low == high:
            return ordered[low]
        weight = rank - low
        return ordered[low] * (1 - weight) + ordered[high] * weight


    def summarize(values: List[float]) -> Dict[str, float]:
        if not values:
            return {
                "avg": math.nan,
                "min": math.nan,
                "max": math.nan,
                "p50": math.nan,
                "p95": math.nan,
                "p99": math.nan,
            }
        return {
            "avg": sum(values) / len(values),
            "min": min(values),
            "max": max(values),
            "p50": percentile(values, 50),
            "p95": percentile(values, 95),
            "p99": percentile(values, 99),
        }

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.legend import Legend
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required. Install it with: pip install openpyxl"
    ) from exc


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
SUMMARY_METRIC_KEYS = ["exec_avg", "exec_min", "exec_max", "exec_p50", "exec_p95", "exec_p99"]
SUMMARY_TABLE_HEADERS = [
    "variant",
    "count",
    "exec_avg",
    "exec_min",
    "exec_max",
    "exec_p50",
    "exec_p95",
    "exec_p99",
]
DETAIL_INTEGER_COLS = {"party_count", "service_count"}
DETAIL_EXTRA_NUMERIC_COLS = {"shared_read", "shared_hit", "shared_dirtied"} | DETAIL_INTEGER_COLS


def die(message: str) -> None:
    print(f"error: {message}", file=sys.stderr)
    raise SystemExit(1)


def parse_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def parse_round_summaries(summary_path: Path) -> List[tuple[int, List[Dict[str, str]]]]:
    round_summaries: List[tuple[int, List[Dict[str, str]]]] = []
    pattern = re.compile(r"^summary-round(\d+)\.csv$")
    for path in sorted(summary_path.parent.glob("summary-round*.csv")):
        match = pattern.match(path.name)
        if not match:
            continue
        round_number = int(match.group(1))
        round_summaries.append((round_number, parse_csv(path)))
    return sorted(round_summaries, key=lambda item: item[0])


def parse_optional_float(value: Optional[str]) -> Optional[float]:
    if value in (None, "", "None"):
        return None
    try:
        return float(value)
    except ValueError:
        return None


def autosize_columns(ws) -> None:
    for col_index in range(1, ws.max_column + 1):
        max_len = 0
        col = get_column_letter(col_index)
        for row_index in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_index, column=col_index)
            val = cell.value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col].width = min(max_len + 2, 60)


def build_details_sheet(
    wb: Workbook,
    sheet_name: str,
    rows: List[Dict[str, str]],
    headers: List[str],
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)

    for col_index in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).coordinate}"

    numeric_cols = [
        h
        for h in headers
        if h.startswith("exec_")
        or h.startswith("read_")
        or h.startswith("hit_")
        or h in DETAIL_EXTRA_NUMERIC_COLS
    ]
    for col_index, header in enumerate(headers, start=1):
        if header in numeric_cols:
            for row_index in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_index, column=col_index)
                if cell.value == "":
                    continue
                try:
                    cell.value = float(cell.value)
                except ValueError:
                    continue
            for row_index in range(2, ws.max_row + 1):
                fmt = "0" if header in DETAIL_INTEGER_COLS else "0.00"
                ws.cell(row=row_index, column=col_index).number_format = fmt

    if ws.max_row > 1:
        for header in numeric_cols:
            if header in DETAIL_INTEGER_COLS:
                continue
            col_idx = headers.index(header) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="63BE7B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="F8696B",
                ),
            )

    autosize_columns(ws)


def summarize_rows_by_variant(rows: List[Dict[str, str]]) -> List[List[Union[float, int, str]]]:
    grouped: Dict[str, Dict[str, List[float]]] = defaultdict(
        lambda: {key: [] for key in SUMMARY_METRIC_KEYS}
    )
    for row in rows:
        variant = row.get("variant") or ""
        if not variant:
            continue
        for key in SUMMARY_METRIC_KEYS:
            parsed = parse_optional_float(row.get(key))
            if parsed is not None:
                grouped[variant][key].append(parsed)

    table_rows: List[List[Union[float, int, str]]] = []
    for variant in sorted(grouped.keys()):
        bucket = grouped[variant]
        stats = {
            "exec_avg": summarize(bucket["exec_avg"])["avg"],
            "exec_min": summarize(bucket["exec_min"])["min"],
            "exec_max": summarize(bucket["exec_max"])["max"],
            "exec_p50": summarize(bucket["exec_p50"])["avg"],
            "exec_p95": summarize(bucket["exec_p95"])["avg"],
            "exec_p99": summarize(bucket["exec_p99"])["avg"],
        }
        count = max(
            len(bucket["exec_avg"]),
            len(bucket["exec_p50"]),
            len(bucket["exec_p95"]),
            len(bucket["exec_p99"]),
        )
        table_rows.append(
            [
                variant,
                count,
                stats["exec_avg"],
                stats["exec_min"],
                stats["exec_max"],
                stats["exec_p50"],
                stats["exec_p95"],
                stats["exec_p99"],
            ]
        )
    return table_rows


def write_summary_section(ws, start_row: int, title: str, rows: List[Dict[str, str]]) -> int:
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)

    header_row = start_row + 1
    for col_index, header in enumerate(SUMMARY_TABLE_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    data_start = header_row + 1
    summary_rows = summarize_rows_by_variant(rows)
    if summary_rows:
        for offset, row_values in enumerate(summary_rows):
            for col_index, value in enumerate(row_values, start=1):
                ws.cell(row=data_start + offset, column=col_index, value=value)

    end_row = data_start + len(summary_rows) - 1
    if summary_rows:
        for col_index in range(3, len(SUMMARY_TABLE_HEADERS) + 1):
            for row_index in range(data_start, end_row + 1):
                cell = ws.cell(row=row_index, column=col_index)
                cell.number_format = "0.00"

        for col_idx in range(3, 9):
            col_letter = ws.cell(row=header_row, column=col_idx).column_letter
            ws.conditional_formatting.add(
                f"{col_letter}{data_start}:{col_letter}{end_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="63BE7B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="F8696B",
                ),
            )

        chart = BarChart()
        chart.type = "bar"
        chart.title = f"{title}: Exec p50/p95/p99 by Variant"
        chart.y_axis.title = "variant"
        chart.x_axis.title = "ms"
        chart.legend = Legend()
        data = Reference(ws, min_col=6, max_col=8, min_row=header_row, max_row=end_row)
        cats = Reference(ws, min_col=1, min_row=data_start, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 14
        ws.add_chart(chart, f"K{start_row}")
    else:
        ws.cell(row=data_start, column=1, value="No data")
        end_row = data_start

    return max(end_row + 3, start_row + 18)


def build_summary_sheet(
    wb: Workbook,
    total_rows: List[Dict[str, str]],
    round_rows: List[tuple[int, List[Dict[str, str]]]],
) -> None:
    ws = wb.active
    ws.title = "Summary"

    next_row = 1
    next_row = write_summary_section(ws, next_row, "Total Aggregate", total_rows)
    for round_number, rows in round_rows:
        next_row = write_summary_section(ws, next_row, f"Round {round_number} Aggregate", rows)

    ws.freeze_panes = "A3"
    for col_index in range(3, len(SUMMARY_TABLE_HEADERS) + 1):
        for row_index in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_index, column=col_index)
            if isinstance(cell.value, (float, int)):
                cell.number_format = "0.00"

    autosize_columns(ws)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate Excel summary from summary.csv")
    parser.add_argument("summary_csv", help="Path to summary.csv")
    parser.add_argument("--out", help="Output .xlsx path (default: alongside input)")
    args = parser.parse_args()

    summary_path = Path(args.summary_csv)
    if not summary_path.is_file():
        die(f"summary.csv not found: {summary_path}")

    out_path = Path(args.out) if args.out else summary_path.with_suffix(".xlsx")

    rows = parse_csv(summary_path)
    if not rows:
        die("summary.csv has no rows")
    round_summaries = parse_round_summaries(summary_path)

    wb = Workbook()

    build_summary_sheet(wb, rows, round_summaries)
    details_headers = list(rows[0].keys())
    build_details_sheet(wb, "Details - total", rows, details_headers)
    for round_number, round_rows in round_summaries:
        build_details_sheet(wb, f"Details - r{round_number}", round_rows, details_headers)

    wb.save(out_path)
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
